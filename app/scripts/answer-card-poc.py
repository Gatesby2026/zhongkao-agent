#!/usr/bin/env python3

import argparse
import json
import re
import subprocess
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional, Union

import numpy as np
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageOps


SUBJECT_CONFIG = {
    "math": {
        "default_dir": "knowledge-original/初三一模-数学",
        "score_file": "数学一模小分.xlsx",
        "choice_page": "IMG_1908.HEIC",
        "choice_questions": list(range(1, 9)),
        "choice_option_centers": {
            1: [(690, 2220), (835, 2220), (978, 2220), (1120, 2220)],
            2: [(1785, 2220), (1928, 2220), (2070, 2220), (2212, 2220)],
            3: [(2810, 2220), (2952, 2220), (3095, 2220), (3238, 2220)],
            4: [(3500, 2220), (3642, 2220), (3785, 2220), (3928, 2220)],
            5: [(690, 2440), (835, 2440), (978, 2440), (1120, 2440)],
            6: [(1785, 2440), (1928, 2440), (2070, 2440), (2212, 2440)],
            7: [(2810, 2440), (2952, 2440), (3095, 2440), (3238, 2440)],
            8: [(3500, 2440), (3642, 2440), (3785, 2440), (3928, 2440)],
        },
        "fill_rois": {
            "9": (430, 3120, 1520, 3265),
            "10": (1680, 3120, 2850, 3265),
            "11": (3050, 3120, 4200, 3265),
            "12": (430, 3300, 1520, 3465),
            "13": (1680, 3300, 2850, 3465),
            "14": (3050, 3300, 4200, 3465),
            "15": (430, 3480, 1520, 3650),
            "16": (1680, 3480, 4200, 3650),
        },
        "fill_answer_regions": [
            {"id": "q09", "item": "9", "label": "9", "box": (650, 2655, 1350, 2825)},
            {"id": "q10", "item": "10", "label": "10", "box": (1900, 2635, 2590, 2825)},
            {"id": "q11", "item": "11", "label": "11", "box": (3180, 2635, 3860, 2825)},
            {"id": "q12-x", "item": "12", "label": "12-x", "box": (650, 2840, 930, 3025)},
            {"id": "q12-y", "item": "12", "label": "12-y", "box": (1180, 2825, 1540, 3025)},
            {"id": "q13", "item": "13", "label": "13", "box": (1830, 2815, 2590, 2970)},
            {"id": "q14", "item": "14", "label": "14", "box": (3180, 2825, 3900, 2985)},
            {"id": "q15", "item": "15", "label": "15", "box": (540, 2985, 1280, 3165)},
            {"id": "q16-1", "item": "16", "label": "16-(1)", "box": (2180, 2980, 2880, 3165)},
            {"id": "q16-2", "item": "16", "label": "16-(2)", "box": (3370, 2975, 3910, 3165)},
        ],
        "solution_rois": {
            "17": (360, 3920, 4200, 4500),
            "18": (360, 4520, 4200, 5650),
        },
    },
    "chinese": {
        "default_dir": "knowledge-original/初三一模-语文",
        "score_file": "一模。语文。小分.xlsx",
    },
}


@dataclass
class ScoreItem:
    item: str
    max_score: Optional[float]
    score: Optional[Union[float, str]]
    description: str


def parse_args():
    parser = argparse.ArgumentParser(description="POC answer-card preprocessing and structured extraction.")
    parser.add_argument("--subject", choices=sorted(SUBJECT_CONFIG), required=True)
    parser.add_argument("--input-dir")
    parser.add_argument("--out-dir", default="data/answer-card-poc")
    return parser.parse_args()


def now_iso():
    return datetime.now(timezone.utc).isoformat()


def write_json(path, payload):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def score_value(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if text in {"", "-"}:
        return text or None
    try:
        return float(text)
    except ValueError:
        return text


def parse_score_label(label):
    text = str(label or "").strip()
    match = re.match(r"([^()]+)\(([\d.]+)_0\)(.*)", text)
    if not match:
        return text, None, ""
    return match.group(1), float(match.group(2)), match.group(3).strip()


def load_scores(score_path):
    workbook = load_workbook(score_path, data_only=True, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    student_name = None
    total = None
    items = []
    for row in rows:
        label, value = row[0], row[1] if len(row) > 1 else None
        if isinstance(label, str) and "考试成绩单" in label:
            student_name = label.strip("“”").replace("的考试成绩单", "")
            continue
        if label == "学科":
            continue
        item, max_score, description = parse_score_label(label)
        if item.startswith("总分"):
            total = score_value(value)
            continue
        if item:
            items.append(ScoreItem(item=item, max_score=max_score, score=score_value(value), description=description))
    return {
        "student_name": student_name,
        "total_score": total,
        "items": [item.__dict__ for item in items],
    }


def convert_heic(input_dir, out_dir):
    out_dir.mkdir(parents=True, exist_ok=True)
    converted = []
    for src in sorted(Path(input_dir).glob("*.HEIC")):
        dst = out_dir / f"{src.stem}.jpg"
        if not dst.exists():
            subprocess.run(["sips", "-s", "format", "jpeg", str(src), "--out", str(dst)], check=True, capture_output=True)
        converted.append({"source": str(src), "image": str(dst), "stem": src.stem})
    return converted


def dark_ratio(image, box):
    crop = image.crop(box).convert("L")
    arr = np.asarray(crop)
    return float((arr < 105).mean())


def detect_math_choices(image, config, out_dir):
    labels = ["A", "B", "C", "D"]
    results = []
    debug = image.copy()
    draw = ImageDraw.Draw(debug)
    crop_dir = out_dir / "crops" / "choice"
    crop_dir.mkdir(parents=True, exist_ok=True)

    for number, centers in config["choice_option_centers"].items():
        option_scores = []
        q_boxes = []
        for label, (cx, cy) in zip(labels, centers):
            box = (cx - 48, cy - 42, cx + 48, cy + 42)
            ratio = dark_ratio(image, box)
            option_scores.append({"label": label, "dark_ratio": ratio, "box": box})
            q_boxes.append(box)
            draw.rectangle(box, outline="red" if ratio > 0.18 else "blue", width=5)
            draw.text((box[0], box[1] - 36), f"{number}{label}:{ratio:.2f}", fill="red" if ratio > 0.18 else "blue")
            image.crop(box).save(crop_dir / f"q{number:02d}-{label}.jpg")
        selected = max(option_scores, key=lambda item: item["dark_ratio"])
        sorted_scores = sorted(option_scores, key=lambda item: item["dark_ratio"], reverse=True)
        confidence = sorted_scores[0]["dark_ratio"] - sorted_scores[1]["dark_ratio"]
        results.append({
            "number": number,
            "type": "choice",
            "recognized_answer": selected["label"],
            "confidence": round(confidence, 4),
            "option_scores": [{k: v for k, v in item.items() if k != "box"} for item in option_scores],
            "needs_review": confidence < 0.05 or selected["dark_ratio"] < 0.12,
        })
    debug.save(out_dir / "math-choice-debug.jpg")
    return results


def crop_rois(image, rois, out_dir, prefix):
    crop_dir = out_dir / "crops" / prefix
    crop_dir.mkdir(parents=True, exist_ok=True)
    results = []
    debug = image.copy()
    draw = ImageDraw.Draw(debug)
    for key, box in rois.items():
        path = crop_dir / f"q{key}.jpg"
        image.crop(box).save(path)
        draw.rectangle(box, outline="green", width=5)
        draw.text((box[0], box[1] - 42), f"q{key}", fill="green")
        results.append({"number": key, "type": prefix, "crop": str(path), "ocr_status": "pending"})
    debug.save(out_dir / f"math-{prefix}-debug.jpg")
    return results


def crop_regions(image, regions, out_dir, prefix):
    crop_dir = out_dir / "crops" / prefix
    crop_dir.mkdir(parents=True, exist_ok=True)
    results = []
    debug = image.copy()
    draw = ImageDraw.Draw(debug)
    for region in regions:
        box = tuple(region["box"])
        path = crop_dir / f"{region['id']}.jpg"
        image.crop(box).save(path)
        draw.rectangle(box, outline="orange", width=5)
        draw.text((box[0], max(0, box[1] - 42)), region["label"], fill="orange")
        results.append({
            "id": region["id"],
            "number": region["item"],
            "label": region["label"],
            "type": prefix,
            "box": box,
            "crop": str(path),
            "ocr_status": "pending",
        })
    debug.save(out_dir / f"math-{prefix}-debug.jpg")
    return results


def normalize_answer_crop(crop):
    arr = np.asarray(crop.convert("RGB")).copy()
    gray = np.asarray(crop.convert("L"))
    dark = gray < 145
    rule_rows = []
    min_run = max(80, int(arr.shape[1] * 0.45))
    for row, values in enumerate(dark):
        padded = np.concatenate(([False], values, [False]))
        changes = np.flatnonzero(padded[1:] != padded[:-1])
        if len(changes) < 2:
            continue
        runs = changes[1::2] - changes[::2]
        if len(runs) and int(runs.max()) >= min_run:
            rule_rows.append(row)
    for row in rule_rows:
        y1 = max(0, row - 3)
        y2 = min(arr.shape[0], row + 4)
        arr[y1:y2, :, :] = 255
    cleaned = Image.fromarray(arr)
    padded = ImageOps.expand(cleaned, border=24, fill="white")
    return padded.resize((padded.width * 2, padded.height * 2), Image.Resampling.LANCZOS)


def crop_normalized_answer_regions(image, regions, out_dir):
    crop_dir = out_dir / "crops" / "fill-answer-clean"
    crop_dir.mkdir(parents=True, exist_ok=True)
    results = []
    for region in regions:
        box = tuple(region["box"])
        path = crop_dir / f"{region['id']}.jpg"
        normalize_answer_crop(image.crop(box)).save(path, quality=95)
        results.append({
            "id": region["id"],
            "number": region["item"],
            "label": region["label"],
            "type": "fill-answer-clean",
            "source_box": box,
            "crop": str(path),
            "ocr_status": "pending",
            "normalization": {
                "removed_horizontal_rules": True,
                "padding_px": 24,
                "scale": 2,
            },
        })
    return results


def run_math_poc(input_dir, out_root, scores):
    config = SUBJECT_CONFIG["math"]
    images = convert_heic(input_dir, out_root / "images")
    page = next(item for item in images if item["stem"] == Path(config["choice_page"]).stem)
    image = ImageOps.exif_transpose(Image.open(page["image"])).convert("RGB")
    choices = detect_math_choices(image, config, out_root)
    fills = crop_rois(image, config["fill_rois"], out_root, "fill")
    fill_answers = crop_regions(image, config["fill_answer_regions"], out_root, "fill-answer")
    fill_answer_clean = crop_normalized_answer_regions(image, config["fill_answer_regions"], out_root)
    solutions = crop_rois(image, config["solution_rois"], out_root, "solution")
    return {
        "subject": "math",
        "source_images": images,
        "student": {"name": scores["student_name"]},
        "scores": scores,
        "extraction": {
            "choice": choices,
            "fill": fills,
            "fill_answer": fill_answers,
            "fill_answer_clean": fill_answer_clean,
            "solution": solutions,
        },
        "notes": [
            "Choice answers use local black-pixel detection over fixed template option boxes.",
            "Fill regions keep the old broad template crop for visual audit.",
            "Fill-answer regions are tight handwritten-answer crops for handwriting OCR and downstream analysis.",
            "Fill-answer-clean regions remove printed horizontal rules, add white padding, and upscale for OCR input.",
            "This POC covers math answer card page IMG_1908 only.",
        ],
    }


def run_generic(subject, input_dir, out_root, scores):
    images = convert_heic(input_dir, out_root / "images")
    return {
        "subject": subject,
        "source_images": images,
        "student": {"name": scores["student_name"]},
        "scores": scores,
        "notes": ["Image conversion and score parsing only; subject template is not implemented yet."],
    }


def main():
    args = parse_args()
    config = SUBJECT_CONFIG[args.subject]
    input_dir = Path(args.input_dir or config["default_dir"])
    out_root = Path(args.out_dir) / args.subject
    score_path = input_dir / config["score_file"]
    scores = load_scores(score_path)
    payload = run_math_poc(input_dir, out_root, scores) if args.subject == "math" else run_generic(args.subject, input_dir, out_root, scores)
    payload["generated_at"] = now_iso()
    write_json(out_root / "answer-card-poc.json", payload)
    print(json.dumps({"subject": args.subject, "out": str(out_root / "answer-card-poc.json")}, ensure_ascii=False))


if __name__ == "__main__":
    main()
