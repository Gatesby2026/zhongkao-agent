"""小分表 → scores.json 统一解析层。

三类输入入口（共用同一输出 schema）：
  - .xlsx 班小二「学生成绩单」 (主路径) → parse_scores_xlsx
  - .csv  同列样式                       → parse_scores_csv
  - .jpg/.png/.heic 手机截图              → parse_scores_image (qwen-vl-max OCR)

统一输出：
  {"examTotal": {"scored","fullScore"},
   "items":     [{"xlsxQid","scored","fullScore","desc?"}, ...],  # 原条目（含子号）
   "questions": [{"qId":"Q<主号>","scored","fullScore"}, ...],     # 主题号合并（向后兼容）
   "sections":  [],
   "_student_name": "<可选>",
   "_source":    "xlsx|csv|image",
   "_warnings":  ["..."]}                                          # 解析警告，含 examTotal 不一致等

入口 dispatcher: parse_scores(file_path)
"""
from __future__ import annotations

import csv
import json
import os
import re
from pathlib import Path

import openpyxl

# 学科列样式：  总分(70_0)  /  12(2_0)  /  20(4_0)；
# 班小二附加情况：
#   - 括号后追加知识点描述："2(2_0)能理解并合理使用成语"
#   - 主观题拆子小问："20_1(3_0)"、"23_1_1(2_0)"、作文"27_1(40_0)"
# 第 1 组只捕"主题号"，子号(_\d+)* 吞掉；末尾允许描述（去 $）
_PAREN = re.compile(r"^\s*(总分|\d+)(?:_\d+)*\s*[\(（]\s*(\d+)_(\d+)\s*[\)）]")


def _num(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(str(v).strip())
    except ValueError:
        return 0.0


def _full(intpart: str, dec: str) -> float:
    """(70_0)→70.0  (2_5)→2.5"""
    return float(f"{intpart}.{dec}")


def parse_scores_xlsx(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    # 取第一个含「成绩单」的 sheet，否则首个
    sn = next((s for s in wb.sheetnames if "成绩" in s), wb.sheetnames[0])
    ws = wb[sn]

    rows = [[("" if c is None else c) for c in r]
            for r in ws.iter_rows(values_only=True)]
    if not rows:
        raise ValueError("空表")

    # 学生名：r1 形如 “贾小淇”的考试成绩单
    name = ""
    m = re.search(r"[“\"']?\s*([一-龥·]{2,6})\s*[”\"']?\s*的.*成绩单",
                  str(rows[0][0]))
    if m:
        name = m.group(1)

    # 完整 xlsx 题号正则（含子号），用于 items 的 xlsxQid 字段
    _LABEL_QID = re.compile(r"^\s*(总分|\d+(?:_\d+)*)")
    exam_total = {"scored": 0.0, "fullScore": 0.0}
    items: list[dict] = []                   # 保留每行原貌（含子号），供 align
    agg: dict[int, dict] = {}                # 主题号合并（向后兼容 questions）
    for r in rows:
        label = str(r[0]).strip()
        scored = _num(r[1]) if len(r) > 1 else 0.0
        pm = _PAREN.match(label)
        if not pm:
            continue
        key, ip, dp = pm.group(1), pm.group(2), pm.group(3)
        full = _full(ip, dp)
        if key == "总分":
            exam_total = {"scored": _i(scored), "fullScore": _i(full)}
            continue
        # 取完整 xlsx 题号文本（如 "27_1" "23_1_1"，整数题保持 "27"）
        qid_text = _LABEL_QID.match(label).group(1)
        items.append({
            "xlsxQid": qid_text,
            "scored": _i(scored),
            "fullScore": _i(full),
            "desc": label[len(pm.group(0)):].strip(),
        })
        # 兼容旧 questions：主题号合并
        qid_int = int(key)
        cur = agg.setdefault(qid_int, {"scored": 0.0, "fullScore": 0.0})
        cur["scored"] += scored
        cur["fullScore"] += full
    questions = [
        {"qId": f"Q{n}",
         "scored": _i(agg[n]["scored"]),
         "fullScore": _i(agg[n]["fullScore"])}
        for n in sorted(agg)
    ]
    # items 一律按 xlsxQid (主+子号) 自然排序——班小二有时按知识点导出，
    # 行序与卷面题号顺序不一致；align 算法假设同序，必须先 sort
    def _sort_key(it):
        parts = [int(p) for p in str(it["xlsxQid"]).split("_") if p.isdigit()]
        return tuple(parts)
    items.sort(key=_sort_key)

    if not questions:
        raise ValueError("未解析到任何小题分（格式不符？）")

    warnings: list[str] = []
    # examTotal 兜底：缺总分则按小题求和
    if exam_total["fullScore"] == 0:
        exam_total = {
            "scored": _i(sum(q["scored"] for q in questions)),
            "fullScore": _i(sum(q["fullScore"] for q in questions)),
        }
    else:
        # examTotal 与小题求和一致性软校验（容差 5 分；C-选 2）
        diff_full = abs(sum(it["fullScore"] for it in items) - exam_total["fullScore"])
        diff_scored = abs(sum(it["scored"] for it in items) - exam_total["scored"])
        if diff_full > 5:
            warnings.append(
                f"小分表满分合计 {sum(it['fullScore'] for it in items)} 与"
                f"总分行 {exam_total['fullScore']} 差 {diff_full:.1f} 分，"
                "可能漏题或二选一作文双倍累加；模块维度数据按 yaml 题号合并")
        if diff_scored > 5:
            warnings.append(
                f"小分表得分合计 {sum(it['scored'] for it in items)} 与"
                f"总分行 {exam_total['scored']} 差 {diff_scored:.1f} 分，请核对")

    return {
        "examTotal": exam_total,
        "items": items,          # 每行原貌（保 xlsxQid 子号）→ schemas 调 align 与 yaml 题号对齐
        "questions": questions,  # 主题号合并；schemas 在 items 缺失时降级用
        "sections": [],          # 班小二无分段信息；build_report 容忍空
        "_student_name": name,
        "_source": "xlsx",
        "_warnings": warnings,
    }


# ─── CSV 支持（班小二亦可导出 csv，列样式同 xlsx）─────────────────────────
def parse_scores_csv(csv_path: Path) -> dict:
    rows: list[list] = []
    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        for r in csv.reader(f):
            rows.append([("" if c is None else c) for c in r])
    return _parse_rows_to_scores(rows, source="csv")


# ─── 图片 OCR（家长截图班小二/班级群截图）─────────────────────────────────
_IMAGE_PROMPT = """这是一张学生考试小分表（电子表格/纸质打印的截图）。
请逐题抄录：每个小题/子小问的「题号、满分、得分」。

严格输出 JSON（不要 markdown 围栏，不要解释）：
{
  "examTotal": {"scored": <学生总分>, "fullScore": <试卷满分>},
  "studentName": "<学生姓名，若可见，否则空串>",
  "items": [
    {"xlsxQid": "1",    "fullScore": 2.0, "scored": 2.0, "desc": "<可选知识点描述>"},
    {"xlsxQid": "27_1", "fullScore": 40.0,"scored": 34.0, "desc": "作文"}
  ]
}

要求：
- xlsxQid 原文抄录（带下划线子号如 "27_1"、"23_1_1"）
- 满分/得分必须是数字；得分缺空格/横线/'-' 当 0
- examTotal 必须从表格的「总分」或「合计」行抄读，不要自己求和
- 不清楚 / 看不清的题不要瞎猜，宁缺勿错
- 输出 items 必须有内容（≥1 题），否则任务失败"""


def parse_scores_image(image_path: Path) -> dict:
    """qwen-vl-max 看小分截图 → 同 schema。

    失败抛 ValueError，调用方（/scores endpoint）返回 422 让前端提示
    "截图识别不准，请换 xlsx 或转 AI 自动判分"（口径 A-选 1）。
    """
    try:
        from PIL import Image, ImageOps
        try:
            import pillow_heif
            pillow_heif.register_heif_opener()
        except Exception:
            pass
    except ImportError:
        raise ValueError("缺 Pillow（图片识别必需）")
    import base64
    import io
    try:
        import openai
    except ImportError:
        raise ValueError("缺 openai 包")
    key = os.environ.get("DASHSCOPE_API_KEY")
    if not key:
        raise ValueError("缺 DASHSCOPE_API_KEY 环境变量")

    # 1. 正立化 + 缩放（与 imgnorm 同口径）——避免手机 EXIF 横置
    im = Image.open(image_path)
    im = ImageOps.exif_transpose(im)
    if im.mode != "RGB":
        im = im.convert("RGB")
    w, h = im.size
    max_dim = 2400
    if max(w, h) > max_dim:
        s = max_dim / float(max(w, h))
        im = im.resize((round(w * s), round(h * s)), Image.LANCZOS)
    buf = io.BytesIO()
    im.save(buf, "JPEG", quality=92)
    b64 = base64.b64encode(buf.getvalue()).decode()

    # 2. 调 qwen-vl-max
    client = openai.OpenAI(
        api_key=key,
        base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",
    )
    resp = client.chat.completions.create(
        model="qwen-vl-max",
        messages=[{"role": "user", "content": [
            {"type": "image_url",
             "image_url": {"url": f"data:image/jpeg;base64,{b64}"}},
            {"type": "text", "text": _IMAGE_PROMPT},
        ]}],
        temperature=0.0, max_tokens=4096,
        response_format={"type": "json_object"}, timeout=120,
    )
    raw = resp.choices[0].message.content.strip()
    raw = re.sub(r"^```(?:json)?\s*|\s*```$", "", raw, flags=re.S).strip()
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        raise ValueError(f"截图识别返回非 JSON：{e}")

    items_raw = data.get("items") or []
    if not isinstance(items_raw, list) or not items_raw:
        raise ValueError("截图未识别到任何小题（请确保拍到完整小分表）")

    # 3. 规整 + 转同 schema
    items: list[dict] = []
    for it in items_raw:
        if not isinstance(it, dict): continue
        qid = str(it.get("xlsxQid") or it.get("qId") or "").strip()
        if not qid: continue
        m = re.match(r"^(?:Q\s*)?(\d+(?:_\d+)*)", qid)
        if not m: continue
        items.append({
            "xlsxQid": m.group(1),
            "scored": _i(_num(it.get("scored"))),
            "fullScore": _i(_num(it.get("fullScore"))),
            "desc": str(it.get("desc") or "").strip(),
        })
    if not items:
        raise ValueError("截图条目均无法解析（题号/分数缺）")

    # 主题号合并 questions
    agg: dict[int, dict] = {}
    for it in items:
        try:
            qn = int(it["xlsxQid"].split("_")[0])
        except (ValueError, KeyError):
            continue
        cur = agg.setdefault(qn, {"scored": 0.0, "fullScore": 0.0})
        cur["scored"] += it["scored"]
        cur["fullScore"] += it["fullScore"]
    questions = [{"qId": f"Q{n}", "scored": _i(agg[n]["scored"]),
                  "fullScore": _i(agg[n]["fullScore"])}
                 for n in sorted(agg)]

    # examTotal
    et = data.get("examTotal") or {}
    exam_total = {
        "scored": _i(_num(et.get("scored"))),
        "fullScore": _i(_num(et.get("fullScore"))),
    }
    warnings: list[str] = []
    if exam_total["fullScore"] == 0:
        exam_total = {
            "scored": _i(sum(q["scored"] for q in questions)),
            "fullScore": _i(sum(q["fullScore"] for q in questions)),
        }
        warnings.append("截图未读到总分行，按小题求和兜底")
    else:
        diff_full = abs(sum(it["fullScore"] for it in items) - exam_total["fullScore"])
        if diff_full > 5:
            warnings.append(
                f"截图总分 {exam_total['fullScore']} 与小题合计 "
                f"{sum(it['fullScore'] for it in items)} 差 {diff_full:.1f} 分；"
                "模块维度按 yaml 对齐时可能受影响")

    # items 排序
    def _sort_key(it):
        parts = [int(p) for p in str(it["xlsxQid"]).split("_") if p.isdigit()]
        return tuple(parts)
    items.sort(key=_sort_key)

    return {
        "examTotal": exam_total,
        "items": items,
        "questions": questions,
        "sections": [],
        "_student_name": str(data.get("studentName") or "").strip(),
        "_source": "image",
        "_warnings": warnings,
    }


# ─── 内部：csv 行 → scores（与 xlsx 共用核心解析）─────────────────────────
def _parse_rows_to_scores(rows: list[list], source: str) -> dict:
    """xlsx / csv 行级解析（接受 ws 或 csv reader 抽好的二维 list）。"""
    if not rows:
        raise ValueError("空表")
    name = ""
    m = re.search(r"[“\"']?\s*([一-龥·]{2,6})\s*[”\"']?\s*的.*成绩单",
                  str(rows[0][0]) if rows[0] else "")
    if m:
        name = m.group(1)
    _LABEL_QID = re.compile(r"^\s*(总分|\d+(?:_\d+)*)")
    exam_total = {"scored": 0.0, "fullScore": 0.0}
    items: list[dict] = []
    agg: dict[int, dict] = {}
    for r in rows:
        if not r: continue
        label = str(r[0]).strip()
        scored = _num(r[1]) if len(r) > 1 else 0.0
        pm = _PAREN.match(label)
        if not pm: continue
        key, ip, dp = pm.group(1), pm.group(2), pm.group(3)
        full = _full(ip, dp)
        if key == "总分":
            exam_total = {"scored": _i(scored), "fullScore": _i(full)}
            continue
        qid_text = _LABEL_QID.match(label).group(1)
        items.append({"xlsxQid": qid_text, "scored": _i(scored),
                      "fullScore": _i(full),
                      "desc": label[len(pm.group(0)):].strip()})
        qid_int = int(key)
        cur = agg.setdefault(qid_int, {"scored": 0.0, "fullScore": 0.0})
        cur["scored"] += scored
        cur["fullScore"] += full
    if not items:
        raise ValueError("未解析到任何小题分（格式不符？）")
    questions = [{"qId": f"Q{n}", "scored": _i(agg[n]["scored"]),
                  "fullScore": _i(agg[n]["fullScore"])}
                 for n in sorted(agg)]

    def _sort_key(it):
        parts = [int(p) for p in str(it["xlsxQid"]).split("_") if p.isdigit()]
        return tuple(parts)
    items.sort(key=_sort_key)

    warnings: list[str] = []
    if exam_total["fullScore"] == 0:
        exam_total = {
            "scored": _i(sum(q["scored"] for q in questions)),
            "fullScore": _i(sum(q["fullScore"] for q in questions)),
        }
    else:
        diff_full = abs(sum(it["fullScore"] for it in items) - exam_total["fullScore"])
        if diff_full > 5:
            warnings.append(
                f"满分合计 {sum(it['fullScore'] for it in items)} 与"
                f"总分行 {exam_total['fullScore']} 差 {diff_full:.1f} 分")

    return {
        "examTotal": exam_total, "items": items, "questions": questions,
        "sections": [], "_student_name": name,
        "_source": source, "_warnings": warnings,
    }


# ─── Dispatcher：按扩展名分流（/scores endpoint 入口）────────────────────
def parse_scores(file_path: Path) -> dict:
    suffix = Path(file_path).suffix.lower()
    if suffix in (".xlsx",):
        return parse_scores_xlsx(file_path)
    if suffix in (".xls",):
        raise ValueError(
            "暂不支持老版 .xls，请用 Excel/WPS 另存为 .xlsx 后重试")
    if suffix == ".csv":
        return parse_scores_csv(file_path)
    if suffix in (".jpg", ".jpeg", ".png", ".heic", ".webp"):
        return parse_scores_image(file_path)
    raise ValueError(f"不支持的小分表格式：{suffix}")


def _i(x: float):
    """整数则去 .0"""
    return int(x) if float(x).is_integer() else round(float(x), 2)


def write_scores_json(xlsx_path: Path, out_path: Path) -> dict:
    data = parse_scores_xlsx(xlsx_path)
    out = {k: v for k, v in data.items() if not k.startswith("_")}
    out_path.write_text(json.dumps(out, ensure_ascii=False, indent=2),
                        encoding="utf-8")
    return data


if __name__ == "__main__":
    import sys
    d = parse_scores_xlsx(Path(sys.argv[1]))
    print(json.dumps(d, ensure_ascii=False, indent=2))
